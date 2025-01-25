import csv
from tkinter import *
import datetime
import pymysql
from openpyxl import Workbook, load_workbook
import os 

# Create a single database to hold both tables
def create_besant_technology_database():
    connection = pymysql.connect(host="localhost", user="root", passwd="")
    cursor = connection.cursor()
    cursor.execute("CREATE DATABASE IF NOT EXISTS besant_technology_db")  # Single database
    cursor.execute("USE besant_technology_db")  # Use the single database
    cursor.execute('''CREATE TABLE IF NOT EXISTS Besant_enquiry_form
                   (date text, name text, mobile text, alternate_no text, email text, address text, course_interested text, 
                   batch_preference text, how_you_came_to_know_us text, are_you_experienced_or_fresher text, 
                   contact_person_from_besant_technology text, counselor text, fee text, comments text)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS Besant_registration_form
                   (date text, name text, mobile text, alternate_no text, email text, address text, course_interested text, 
                   batch_preference text, how_you_came_to_know_us text, are_you_experienced_or_fresher text, 
                   contact_person_from_besant_technology text, counselor text, fee text, comments text)''')
    
    cursor.close()
    connection.close()

# Create the databases and tables
create_besant_technology_database()

# Create the main window using Tkinter
window = Tk()
window.title("BESANT TECHNOLOGY")  # Set the title of the window
window.configure(bg="light blue")  # Set the background color of the window
window.geometry("1000x1000")  # Set the size of the window

# Get the current date
today = datetime.date.today()

# Create labels and entries directly in the window
Label(window, text="BESANT TECHNOLOGY\nENQUIRY FORM", fg="red", font=("Arial", 12, "bold")).grid(row=0, column=1)

# Date field
Label(window, text="Date:", bg="light blue").grid(row=1, column=0, sticky="w")
eDate = Entry(window, width=50)
eDate.insert(0, today.strftime("%d/%m/%Y"))  # Set the current date as the default value
eDate.grid(row=1, column=1)

# Create the rest of the labels and entries in the same way for each field
Label(window, text="NAME:", bg="light blue").grid(row=2, column=0, sticky="w")
eName = Entry(window, width=50)
eName.grid(row=2, column=1)

Label(window, text="MOBILE NO:", bg="light blue").grid(row=3, column=0, sticky="w")
eMobile = Entry(window, width=50)
eMobile.grid(row=3, column=1)

Label(window, text="ALTERNATE NO:", bg="light blue").grid(row=4, column=0, sticky="w")
eAlt = Entry(window, width=50)
eAlt.grid(row=4, column=1)

Label(window, text="EMAIL ID:", bg="light blue").grid(row=5, column=0, sticky="w")
eEmail = Entry(window, width=50)
eEmail.grid(row=5, column=1)

Label(window, text="ADDRESS:", bg="light blue").grid(row=6, column=0, sticky="w")
eAddress = Entry(window, width=50)
eAddress.grid(row=6, column=1)

Label(window, text="COURSE INTERESTED:", bg="light blue").grid(row=7, column=0, sticky="w")
eCourse = Entry(window, width=50)
eCourse.grid(row=7, column=1)

Label(window, text="BATCH PREFERENCE:", bg="light blue").grid(row=8, column=0, sticky="w")
eBatch = Entry(window, width=50)
eBatch.grid(row=8, column=1)

Label(window, text="How You Came To Know Us:", bg="light blue").grid(row=9, column=0, sticky="w")
eHow = Entry(window, width=50)
eHow.grid(row=9, column=1)

Label(window, text="Are You Experienced or Fresher:", bg="light blue").grid(row=10, column=0, sticky="w")
eExperiance = Entry(window, width=50)
eExperiance.grid(row=10, column=1)

Label(window, text="Contact Person From Besant Technology:", bg="light blue").grid(row=11, column=0)
eContact = Entry(window, width=50)
eContact.grid(row=11, column=1)

Label(window, text="Counselor:", bg="light blue").grid(row=12, column=0, sticky="w")
eCounselor = Entry(window, width=50)
eCounselor.grid(row=12, column=1)

Label(window, text="Fee:", bg="light blue").grid(row=13, column=0, sticky="w")
eFee = Entry(window, width=50)
eFee.grid(row=13, column=1)

Label(window, text="Comments:", bg="light blue").grid(row=14, column=0, sticky="w")
eComments = Entry(window, width=50)
eComments.grid(row=14, column=1)

# Create checkboxes for "Enquiry" and "Registration"
var1 = IntVar()  # Define an integer variable for the "Enquiry" checkbox
Checkbutton(window, text="Enquiry", variable=var1).grid(row=15, column=1, sticky="w")

var2 = IntVar()  # Define an integer variable for the "Registration" checkbox
Checkbutton(window, text="Registration", variable=var2).grid(row=15, column=1, sticky="n")

# Function to save data to the MySQL database
def save_to_db():
    data = (eDate.get(), eName.get(), eMobile.get(), eAlt.get(), eEmail.get(), eAddress.get(), eCourse.get(),
            eBatch.get(), eHow.get(), eExperiance.get(), eContact.get(), eCounselor.get(), eFee.get(),
            eComments.get())

    # Ensure you are connecting to the correct database
    connection = pymysql.connect(host="localhost", user="root", passwd="", database="besant_technology_db")
    cursor = connection.cursor()

    if var1.get() == 1:  # Enquiry checkbox is selected
        cursor.execute(
            "INSERT INTO Besant_enquiry_form VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", data)
        connection.commit()
        print("Data saved to Enquiry table.")

    if var2.get() == 1:  # Registration checkbox is selected
        cursor.execute(
            "INSERT INTO Besant_registration_form VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            data)
        connection.commit()
        print("Data saved to Registration table.")

    cursor.close()
    connection.close()

# Function to save data to an Excel file
def save_to_excel():
    try:
        filename = 'Besant_enquiry_forms.xlsx'
        if os.path.exists(filename):  # Check if the file already exists
            workbook = load_workbook(filename)  # Load the existing workbook
            worksheet = workbook.active  # Get the active worksheet
        else:
            workbook = Workbook()  # Create a new workbook if it doesn't exist
            worksheet = workbook.active  # Get the active worksheet
            headers = ['Date', 'Name', 'Mobile', 'Alternate No', 'Email', 'Address', 'Course Interested',
                       'Batch Preference', 'How You Came To Know Us', 'Are You Experienced or Fresher',
                       'Contact Person From Besant Technology', 'Counselor', 'Fee', 'Comments']
            worksheet.append(headers)  # Add headers to the new worksheet

        data = [eDate.get(), eName.get(), eMobile.get(), eAlt.get(), eEmail.get(), eAddress.get(), eCourse.get(),
                eBatch.get(), eHow.get(), eExperiance.get(), eContact.get(), eCounselor.get(), eFee.get(),
                eComments.get()]  # Collect the data from the form

        worksheet.append(data)  # Append the data to the worksheet
        workbook.save(filename)  # Save the workbook
        print("Data saved to Excel file successfully.")  # Print a success message
    except Exception as e:
        print("Error saving to Excel file:", e)  # Print any error that occurs

# Function to handle form submission
def submit_form():
    print("Form submitted")
    save_to_db()  # This will now save to one or both tables in the same database based on checkbox selection
    save_to_excel()
    # Clear the form fields after submission
    eDate.delete(0, END)
    eName.delete(0, END)
    eMobile.delete(0, END)
    eAlt.delete(0, END)
    eEmail.delete(0, END)
    eAddress.delete(0, END)
    eCourse.delete(0, END)
    eBatch.delete(0, END)
    eHow.delete(0, END)
    eExperiance.delete(0, END)
    eContact.delete(0, END)
    eCounselor.delete(0, END)
    eFee.delete(0, END)
    eComments.delete(0, END)

# Function to handle form cancellation
def cancel_form():
    print("Form cancelled")  # Print a cancellation message
    # Clear the form fields
    eDate.delete(0, END)
    eName.delete(0, END)
    eMobile.delete(0, END)
    eAlt.delete(0, END)
    eEmail.delete(0, END)
    eAddress.delete(0, END)
    eCourse.delete(0, END)
    eBatch.delete(0, END)
    eHow.delete(0, END)
    eExperiance.delete(0, END)
    eContact.delete(0, END)
    eCounselor.delete(0, END)
    eFee.delete(0, END)
    eComments.delete(0, END)
    # Close the MySQL connection and destroy the window
    window.destroy()

# Create buttons for "Submit" and "Cancel"
Button(window, text="Submit", bg="GREEN", command=submit_form).grid(row=16, column=1, sticky="w")
Button(window, text="Cancel", bg="red", command=cancel_form).grid(row=16, column=1, sticky="n")

# Start the Tkinter event loop
window.mainloop()